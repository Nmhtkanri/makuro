# -*- coding: utf-8 -*-
"""
2月控除分のK列と勤務報告書の3月シートDE54(結合セル)の値を突合チェックするスクリプト

使い方:
  python チェックスクリプト.py
"""

import openpyxl
import os
import sys
import csv
import datetime
import win32com.client


# ── 設定 ──────────────────────────────────────────
SOURCE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "控除発生_2月.xlsx")
SHEET_KOUJO = "2月控除分"
SHEET_PATH  = "パスリスト"
TARGET_SHEET = "3月"
CHECK_ROW    = 54
CHECK_COL    = 4   # D列 (D54:E54結合セルの左上)


# ── ユーティリティ ─────────────────────────────────
def normalize_id(val):
    """従業員番号の型ゆれを吸収して文字列化"""
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        try:
            s = str(int(float(s)))
        except ValueError:
            pass
    return s


def normalize_val(val):
    """比較用に値を正規化（数値は小数点以下の誤差を考慮）"""
    if val is None:
        return None
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return val


def run():
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            f"チェック結果_{timestamp}.csv")

    # ── 元データ読み込み (openpyxl: 読み取り専用) ─────
    print(f"元データ読み込み: {SOURCE_FILE}")
    wb_src = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    # 2月控除分シート → {社員番号: K列の値}
    ws_koujo = wb_src[SHEET_KOUJO]
    koujo_k_map = {}
    for row in range(2, ws_koujo.max_row + 1):
        emp_id = normalize_id(ws_koujo.cell(row, 1).value)
        if not emp_id:
            continue
        val = ws_koujo.cell(row, 11).value  # K列
        koujo_k_map[emp_id] = val
    print(f"  2月控除分: {len(koujo_k_map)} 件読み込み")

    # パスリスト → [(従業員番号, パス), ...]
    ws_path = wb_src[SHEET_PATH]
    path_list = []
    for row in range(2, ws_path.max_row + 1):
        emp_id = normalize_id(ws_path.cell(row, 1).value)
        file_path = ws_path.cell(row, 3).value  # C列
        if emp_id:
            path_list.append((emp_id, file_path, row))
    print(f"  パスリスト: {len(path_list)} 件読み込み")

    wb_src.close()

    # ── Excel COM 起動 ────────────────────────────
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    # ── チェック処理 ──────────────────────────────
    results = []  # (従業員番号, パス, 判定, 2月控除分K列値, 勤務報告書DE54値, 備考)
    ok_count = 0
    ng_count = 0
    skip_count = 0

    for emp_id, fpath, src_row in path_list:
        # 社員番号が2月控除分に存在するか
        if emp_id not in koujo_k_map:
            results.append((emp_id, fpath, "スキップ", "", "", "2月控除分に該当社員番号なし"))
            skip_count += 1
            continue

        k_val = koujo_k_map[emp_id]

        # パス未設定
        if not fpath or str(fpath).strip() == "":
            results.append((emp_id, fpath, "スキップ", k_val, "", "パス未設定"))
            skip_count += 1
            continue

        fpath = str(fpath).strip()

        # ファイル存在チェック
        if not os.path.isfile(fpath):
            results.append((emp_id, fpath, "スキップ", k_val, "", "ファイル不存在"))
            skip_count += 1
            continue

        wb = None
        try:
            wb = excel.Workbooks.Open(os.path.abspath(fpath), ReadOnly=True)

            # 3月シート存在チェック
            sheet_names = [wb.Sheets(i).Name for i in range(1, wb.Sheets.Count + 1)]
            if TARGET_SHEET not in sheet_names:
                results.append((emp_id, fpath, "スキップ", k_val, "", "3月シートなし"))
                skip_count += 1
                wb.Close(SaveChanges=False)
                continue

            ws = wb.Sheets(TARGET_SHEET)
            de54_val = ws.Cells(CHECK_ROW, CHECK_COL).Value
            wb.Close(SaveChanges=False)

            # 比較
            k_norm = normalize_val(k_val)
            de54_norm = normalize_val(de54_val)

            if k_norm == de54_norm:
                results.append((emp_id, fpath, "OK", k_val, de54_val, ""))
                ok_count += 1
            else:
                results.append((emp_id, fpath, "NG", k_val, de54_val, "値が不一致"))
                ng_count += 1

        except Exception as e:
            results.append((emp_id, fpath, "エラー", k_val, "", str(e)))
            skip_count += 1
            if wb is not None:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass

    # ── Excel COM 終了 ────────────────────────────
    excel.Quit()

    # ── CSV出力 ───────────────────────────────────
    with open(log_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["従業員番号", "ファイルパス", "判定", "2月控除分_K列", "勤務報告書_DE54", "備考"])
        for r in results:
            writer.writerow(r)

    # ── サマリ表示 ─────────────────────────────────
    print(f"\n{'='*50}")
    print(f"チェック対象: {len(path_list)} 件")
    print(f"OK:          {ok_count} 件")
    print(f"NG:          {ng_count} 件")
    print(f"スキップ:    {skip_count} 件")
    print(f"結果CSV:     {log_path}")
    print(f"{'='*50}")

    if ng_count > 0:
        print(f"\n--- NG一覧 ({ng_count}件) ---")
        for emp_id, fpath, status, k_val, de54_val, note in results:
            if status == "NG":
                fname = os.path.basename(fpath) if fpath else "(なし)"
                print(f"  {emp_id} | {fname} | K列={k_val} / DE54={de54_val}")

    return ok_count, ng_count, skip_count


if __name__ == "__main__":
    run()
