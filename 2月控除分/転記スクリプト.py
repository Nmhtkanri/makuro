# -*- coding: utf-8 -*-
"""
2月控除分の支給額を各勤務報告書の3月シートJ59セルへ転記するスクリプト

元データの読み込みは openpyxl (読み取り専用)、
転記先 .xlsm の書き込みは win32com (Excel COM) で行い、
図形・マクロ・書式を一切壊さない。

使い方:
  python 転記スクリプト.py          # 全件実行
  python 転記スクリプト.py --test    # テスト（1件のみ）
"""

import openpyxl
import os
import sys
import csv
import datetime

# ── 設定 ──────────────────────────────────────────
SOURCE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "控除発生_2月.xlsx")
SHEET_KOUJO = "2月控除分"
SHEET_PATH  = "パスリスト"
TARGET_SHEET = "3月"
TARGET_ROW   = 59
TARGET_COL   = 10  # J列


# ── ユーティリティ ─────────────────────────────────
def normalize_id(val):
    """従業員番号の型ゆれを吸収して文字列化"""
    if val is None:
        return ""
    s = str(val).strip()
    # 12345.0 → 12345
    if s.endswith(".0"):
        try:
            s = str(int(float(s)))
        except ValueError:
            pass
    return s


def run(test_mode=False):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            f"転記ログ_{timestamp}.csv")

    # ── 元データ読み込み (openpyxl: 読み取り専用) ─────
    print(f"元データ読み込み: {SOURCE_FILE}")
    wb_src = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    # 2月控除分シート → {社員番号: 支給分}
    ws_koujo = wb_src[SHEET_KOUJO]
    koujo_map = {}
    for row in range(2, ws_koujo.max_row + 1):
        emp_id = normalize_id(ws_koujo.cell(row, 1).value)
        if not emp_id:
            continue
        val = ws_koujo.cell(row, 12).value  # L列
        koujo_map[emp_id] = val
    print(f"  2月控除分: {len(koujo_map)} 件読み込み")

    # パスリスト → [(従業員番号, パス), ...]
    ws_path = wb_src[SHEET_PATH]
    path_list = []
    for row in range(2, ws_path.max_row + 1):
        emp_id = normalize_id(ws_path.cell(row, 1).value)
        file_path = ws_path.cell(row, 3).value  # C列
        if emp_id:
            path_list.append((emp_id, file_path, row))
    print(f"  パスリスト: {len(path_list)} 件読み込み")

    wb_src.close()

    # ── Excel COM 起動 ────────────────────────────
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    # ── 転記処理 ──────────────────────────────────
    results = []  # (従業員番号, パス, ステータス, 詳細)
    success = 0
    fail = 0
    skip = 0

    targets = path_list[:1] if test_mode else path_list

    for emp_id, fpath, src_row in targets:
        # 社員番号が2月控除分に存在するか
        if emp_id not in koujo_map:
            results.append((emp_id, fpath, "スキップ", "2月控除分に該当社員番号なし"))
            skip += 1
            continue

        val_to_write = koujo_map[emp_id]

        # パス未設定
        if not fpath or str(fpath).strip() == "":
            results.append((emp_id, fpath, "スキップ", "パス未設定"))
            skip += 1
            continue

        fpath = str(fpath).strip()

        # ファイル存在チェック
        if not os.path.isfile(fpath):
            results.append((emp_id, fpath, "スキップ", "ファイル不存在"))
            skip += 1
            continue

        wb = None
        try:
            wb = excel.Workbooks.Open(os.path.abspath(fpath))

            # 3月シート存在チェック
            sheet_names = [wb.Sheets(i).Name for i in range(1, wb.Sheets.Count + 1)]
            if TARGET_SHEET not in sheet_names:
                results.append((emp_id, fpath, "スキップ", "3月シートなし"))
                skip += 1
                wb.Close(SaveChanges=False)
                continue

            ws = wb.Sheets(TARGET_SHEET)
            old_val = ws.Cells(TARGET_ROW, TARGET_COL).Value
            ws.Cells(TARGET_ROW, TARGET_COL).Value = val_to_write
            wb.Save()
            wb.Close(SaveChanges=False)

            results.append((emp_id, fpath, "成功",
                            f"旧値={old_val} → 新値={val_to_write}"))
            success += 1

        except Exception as e:
            results.append((emp_id, fpath, "失敗", str(e)))
            fail += 1
            if wb is not None:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass

    # ── Excel COM 終了 ────────────────────────────
    excel.Quit()

    # ── ログ出力 ──────────────────────────────────
    with open(log_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["従業員番号", "ファイルパス", "ステータス", "詳細"])
        for r in results:
            writer.writerow(r)

    # ── サマリ表示 ─────────────────────────────────
    mode_label = "テスト（1件）" if test_mode else "全件"
    print(f"\n{'='*50}")
    print(f"実行モード: {mode_label}")
    print(f"処理対象:   {len(targets)} 件")
    print(f"成功:       {success} 件")
    print(f"失敗:       {fail} 件")
    print(f"スキップ:   {skip} 件")
    print(f"ログ出力:   {log_path}")
    print(f"{'='*50}")

    if fail > 0 or skip > 0:
        print("\n--- 失敗/スキップ一覧 ---")
        for emp_id, fpath, status, detail in results:
            if status != "成功":
                fname = os.path.basename(fpath) if fpath else "(なし)"
                print(f"  [{status}] {emp_id} | {fname} | {detail}")

    if success > 0:
        print("\n--- 成功一覧 ---")
        for emp_id, fpath, status, detail in results:
            if status == "成功":
                fname = os.path.basename(fpath) if fpath else "(なし)"
                print(f"  {emp_id} | {fname} | {detail}")

    return success, fail, skip


if __name__ == "__main__":
    test_mode = "--test" in sys.argv
    run(test_mode=test_mode)
