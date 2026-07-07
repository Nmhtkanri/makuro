# -*- coding: utf-8 -*-
"""通勤経路×経費申請の突合チェック
経費チェック2026年6月(1).xlsx の「通勤費」シート（jinjer登録の通勤経路）と
経費利用履歴 Rev5.csv（統合一覧表）の交通費行を突合し、
通勤経路上の移動なのに 通勤定期代/通勤交通費（実費）以外 を選択している行を検出する。
"""
import csv
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BASE = Path(r"Y:\給与明細\R8年\7月\経費確認精査データ")
CSV_PATH = BASE / "経費利用履歴 Rev5.csv"
XLSX_PATH = BASE / "経費チェック2026年6月 (1).xlsx"
OUT_PATH = BASE / "経路突合チェック_202606.xlsx"

COMMUTE_TYPES = {"通勤定期代", "通勤交通費（実費）"}

LINE_PAT = re.compile(r"(線$|線・|ライン|行$|バス$|快速|各停|快特|特急|急行|新幹線|徒歩)")


def norm_station(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s)).strip()
    s = re.sub(r"[（(].*?[）)]", "", s)  # 括弧内の路線名等を除去
    s = s.replace(" ", "").replace("　", "")
    s = re.sub(r"駅$", "", s)
    return s


def stations_from_route_text(text):
    """'A → 路線名 → B → 路線名 → C' 形式から駅名らしきトークンだけ抽出"""
    if not text:
        return []
    toks = [t.strip() for t in re.split(r"→|⇒|〜|~", str(text)) if t.strip()]
    out = []
    for t in toks:
        if LINE_PAT.search(t):
            continue
        n = norm_station(t)
        if n:
            out.append(n)
    return out


# ---------- 1. 通勤費シート読み込み ----------
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb["通勤費"]
commute = {}  # emp_id -> {"stations": set, "routes": [repr], "kikan": [利用交通機関]}
header = None
for row in ws.iter_rows(values_only=True):
    if header is None:
        header = list(row)
        continue
    emp = str(row[0]).strip() if row[0] is not None else ""
    if not emp:
        continue
    dep, arr, via1, via2, route_txt, kikan = row[3], row[4], row[5], row[6], row[7], row[8]
    ent = commute.setdefault(emp, {"stations": set(), "routes": [], "kikan": set()})
    sts = [norm_station(x) for x in (dep, arr, via1, via2)]
    sts = [s for s in sts if s]
    sts += stations_from_route_text(route_txt)
    ent["stations"].update(sts)
    rep = "→".join([str(x) for x in (dep, via1, via2, arr) if x]) or (str(route_txt) if route_txt else "")
    if rep:
        ent["routes"].append(rep + (f"[{kikan}]" if kikan else ""))
    if kikan:
        ent["kikan"].add(str(kikan))
wb.close()

n_with_station = sum(1 for v in commute.values() if v["stations"])
print(f"通勤費シート: 従業員 {len(commute)} 名（駅登録あり {n_with_station} 名）")

# ---------- 2. Rev5 CSV 読み込みと判定 ----------
results = []
counts = {}
with open(CSV_PATH, encoding="cp932", newline="") as f:
    reader = csv.DictReader(f)
    for r in reader:
        kikan = (r.get("交通機関") or "").strip()
        board = norm_station(r.get("乗車場所"))
        alight = norm_station(r.get("降車場所"))
        if not kikan and not (board or alight):
            continue  # 交通費行以外（夜間当番手当等）はスキップ
        emp = (r.get("社員番号") or "").strip()
        ent = commute.get(emp)
        stations = ent["stations"] if ent else set()
        routes = " / ".join(ent["routes"]) if ent else ""

        if not ent or not stations:
            match = "通勤経路登録なし"
        elif board and alight and board in stations and alight in stations:
            match = "経路内"
        elif (board and board in stations) or (alight and alight in stations):
            match = "片側一致"
        else:
            match = "一致なし"

        if match == "経路内":
            verdict = "OK（通勤系を選択）" if kikan in COMMUTE_TYPES else "★要確認（経路内なのに通勤系以外を選択）"
        elif match in ("一致なし", "通勤経路登録なし") and kikan in COMMUTE_TYPES:
            verdict = "△逆要確認（通勤系なのに登録経路と不一致）"
        elif match == "片側一致":
            verdict = "参考（片側のみ一致）"
        else:
            verdict = "OK（経路外）"

        counts[verdict] = counts.get(verdict, 0) + 1
        results.append({
            "社員番号": emp, "氏名": (r.get("氏名") or "").strip(),
            "利用日": r.get("利用日(yyyy/mm/dd)") or "", "交通機関": kikan,
            "内訳": (r.get("内訳") or "").strip(),
            "乗車場所": r.get("乗車場所") or "", "降車場所": r.get("降車場所") or "",
            "経路": r.get("経路") or "", "金額": r.get("合計") or "", "往復": r.get("往復") or "",
            "備考(明細)": r.get("備考(明細)") or "",
            "登録通勤経路": routes, "一致": match, "判定": verdict,
        })

print(f"チェック対象の交通費行: {len(results)} 行")
for k in sorted(counts, key=counts.get, reverse=True):
    print(f"  {k}: {counts[k]}")

# 要確認の内訳（従業員別）
flagged = [x for x in results if x["判定"].startswith("★")]
rev_flagged = [x for x in results if x["判定"].startswith("△")]
by_emp = {}
for x in flagged:
    k = (x["社員番号"], x["氏名"])
    by_emp.setdefault(k, []).append(x)
print(f"\n★要確認: {len(flagged)} 行 / {len(by_emp)} 名")
for (emp, name), rows in sorted(by_emp.items()):
    total = sum(int(float(r['金額'] or 0)) for r in rows)
    print(f"  {emp} {name}: {len(rows)}行 計{total:,}円 例: {rows[0]['乗車場所']}→{rows[0]['降車場所']} [{rows[0]['交通機関']}]")
by_emp2 = {}
for x in rev_flagged:
    by_emp2.setdefault((x["社員番号"], x["氏名"]), []).append(x)
print(f"\n△逆要確認: {len(rev_flagged)} 行 / {len(by_emp2)} 名")

# ---------- 3. Excel 出力 ----------
out = Workbook()
cols = ["社員番号", "氏名", "利用日", "交通機関", "内訳", "乗車場所", "降車場所", "経路",
        "金額", "往復", "備考(明細)", "登録通勤経路", "一致", "判定"]
red = PatternFill("solid", fgColor="FFC7CE")
yellow = PatternFill("solid", fgColor="FFEB9C")

def write_sheet(ws_out, rows):
    ws_out.append(cols)
    for c in ws_out[1]:
        c.font = Font(name="Meiryo UI", bold=True)
    for r in rows:
        ws_out.append([r[c] for c in cols])
        if r["判定"].startswith("★"):
            for c in ws_out[ws_out.max_row]:
                c.fill = red
        elif r["判定"].startswith("△"):
            for c in ws_out[ws_out.max_row]:
                c.fill = yellow
    widths = [9, 12, 10, 18, 14, 12, 12, 40, 8, 6, 20, 40, 12, 34]
    for i, w in enumerate(widths, 1):
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = ws_out.dimensions

ws1 = out.active
ws1.title = "要確認"
write_sheet(ws1, flagged + rev_flagged)
ws2 = out.create_sheet("全交通費行")
write_sheet(ws2, results)
out.save(OUT_PATH)
print(f"\n出力: {OUT_PATH}")
